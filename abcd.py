import os
# Cloud 환경 대비: 필요한 라이브러리 설치 (requirements.txt 쓰면 이 줄은 제거해도 됨)
os.system("pip install -q streamlit reportlab pillow openpyxl")

# -*- coding: utf-8 -*-
# 캐스케이드/환기 기성 청구 양식
# - UI: 한 줄(체크박스 · '항목' 라벨 · 드롭다운) + 아래 사진 업로드/미리보기
# - PDF/Excel: 동일한 3열 카드(이미지 + 캡션 + 테두리), 업로드된 사진까지만 출력(빈칸 없음)

import io, re, unicodedata, uuid
from typing import List, Tuple, Optional

import streamlit as st
from PIL import Image

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image as RLImage
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# ───────────────────────────────
# 페이지/폰트/스타일
# ───────────────────────────────
st.set_page_config(page_title="캐스케이드/환기 기성 청구 양식", layout="wide")

def try_register_font():
    candidates = [
        ("NanumGothic", "NanumGothic.ttf"),
        ("MalgunGothic", "C:\\Windows\\Fonts\\malgun.ttf"),
        ("MalgunGothic", "C:/Windows/Fonts/malgun.ttf"),
    ]
    for name, path in candidates:
        try:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont(name, path))
                return name, True
        except Exception:
            pass
    return "Helvetica", False

BASE_FONT, _ = try_register_font()
ss = getSampleStyleSheet()
styles = {
    "title": ParagraphStyle(name="title", parent=ss["Heading1"], fontName=BASE_FONT,
                            fontSize=18, leading=22, alignment=1, spaceAfter=8),
    "cell": ParagraphStyle(name="cell", parent=ss["Normal"], fontName=BASE_FONT,
                           fontSize=10, leading=13),
    "small_center": ParagraphStyle(name="small_center", parent=ss["Normal"], fontName=BASE_FONT,
                                   fontSize=8.5, leading=11, alignment=1),
}

# ───────────────────────────────
# 유틸
# ───────────────────────────────
def sanitize_filename(name: str) -> str:
    name = unicodedata.normalize("NFKD", name)
    return re.sub(r"[\\/:*?\"<>|]", "_", name).strip().strip(".") or "output"

def enforce_aspect_pad(img: Image.Image, target_ratio: float = 4/3) -> Image.Image:
    """이미지를 target_ratio(4:3)로 흰색 패딩해 캔버스 맞춤."""
    w, h = img.size
    cur_ratio = w / h
    if abs(cur_ratio - target_ratio) < 1e-3:
        return img
    if cur_ratio > target_ratio:  # 가로가 더 김 → 세로 확장
        new_h = int(round(w / target_ratio))
        new_w = w
    else:                         # 세로가 더 김 → 가로 확장
        new_w = int(round(h * target_ratio))
        new_h = h
    canvas = Image.new("RGB", (new_w, new_h), (255, 255, 255))
    canvas.paste(img, ((new_w - w) // 2, (new_h - h) // 2))
    return canvas

def _pil_to_bytesio(img: Image.Image, quality=85) -> io.BytesIO:
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality, optimize=True)
    buf.seek(0)
    return buf

# ───────────────────────────────
# PDF 빌더 (3열 카드, 테두리/제목/주소 유지, 빈칸 없음)
# ───────────────────────────────
def build_pdf(doc_title: str, site_addr: str, items: List[Tuple[str, Image.Image]]) -> bytes:
    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4
    MARGIN = 20
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=MARGIN, bottomMargin=MARGIN,
                            leftMargin=MARGIN, rightMargin=MARGIN,
                            title=doc_title)
    story = []
    # 제목
    story.append(Paragraph(doc_title, styles["title"]))
    story.append(Spacer(1, 4))
    # 현장 주소 표 (테두리 유지)
    meta_tbl = Table(
        [[Paragraph("현장 주소", styles["cell"]), Paragraph(site_addr.strip() or "-", styles["cell"])]],
        colWidths=[80, PAGE_W - 2*MARGIN - 80]
    )
    meta_tbl.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 0.9, colors.black),
        ("INNERGRID", (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 8))

    if not items:
        doc.build(story)
        return buf.getvalue()

    # 카드 크기
    col_count = 3
    usable_width = PAGE_W - 2*MARGIN
    col_width = usable_width / col_count
    ROW_HEIGHT = 200
    CAPTION_HEIGHT = 22
    IMAGE_MAX_H = ROW_HEIGHT - CAPTION_HEIGHT - 8
    IMAGE_MAX_W = col_width - 8

    # 카드 생성 (업로드된 사진만)
    cells = []
    for label, pil_img in items:
        pil_img = enforce_aspect_pad(pil_img, 4/3)
        # 크기 조정(너비 기준)
        target_w = IMAGE_MAX_W
        target_h = target_w * 3 / 4
        if target_h > IMAGE_MAX_H:
            target_h = IMAGE_MAX_H
            target_w = target_h * 4 / 3

        bio = _pil_to_bytesio(pil_img)
        rl_img = RLImage(bio, width=target_w, height=target_h)
        rl_img.hAlign = "CENTER"

        card = Table(
            [[rl_img], [Paragraph(label, styles["small_center"])]],
            colWidths=[col_width],
            rowHeights=[ROW_HEIGHT - CAPTION_HEIGHT, CAPTION_HEIGHT]
        )
        card.setStyle(TableStyle([
            ("BOX", (0,0), (-1,-1), 0.4, colors.grey),
            ("VALIGN", (0,0), (-1,0), "MIDDLE"),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("LEFTPADDING", (0,0), (-1,-1), 2),
            ("RIGHTPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING", (0,0), (-1,-1), 2),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ]))
        cells.append(card)

    # 3열 그리드(빈칸 없이)
    rows = [cells[i:i+3] for i in range(0, len(cells), 3)]
    # 마지막 행이 3개 미만이어도 그대로 출력 (빈칸 안 채움)
    grid_tbl = Table(rows, colWidths=[col_width]*3, rowHeights=[ROW_HEIGHT]*len(rows))
    story.append(grid_tbl)

    doc.build(story)
    return buf.getvalue()

# ───────────────────────────────
# Excel 빌더 (PDF와 동일한 3열 카드, 1페이지 맞춤, 테두리/제목/주소 유지)
# ───────────────────────────────
def build_excel_like_pdf(doc_title: str, site_addr: str, items: List[Tuple[str, Image.Image]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "기성청구"

    # 페이지 설정: A4, 가로, 여백 축소, 한 페이지에 맞추기
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # 제목
    ws.merge_cells("A1:I1")
    ws["A1"] = doc_title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 현장 주소 표 (테두리 유지)
    ws["A3"] = "현장 주소"
    ws.merge_cells("B3:I3")
    ws["B3"] = site_addr or "-"
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A3"].border = border
    for cell in ws["B3:I3"][0]:
        cell.border = border

    # 열 너비/행 높이(대략 3열 카드 배치: A-C, D-F, G-I)
    for col in range(1, 10):  # A..I
        ws.column_dimensions[get_column_letter(col)].width = 14
    # 카드 한 세트(이미지영역+캡션영역) 높이
    CARD_IMG_ROWS = 16
    CAPTION_ROWS = 2

    # 카드 그리드 시작 행
    start_row = 5
    col_groups = [(1,3), (4,6), (7,9)]  # (A-C), (D-F), (G-I)

    # 작성 함수: 병합영역 테두리 적용
    def set_region_border(r1, c1, r2, c2):
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).border = border

    r = start_row
    cgi = 0  # column group index

    for idx, (label, pil_img) in enumerate(items):
        # 열 그룹
        c1, c2 = col_groups[cgi]
        # 이미지 영역: r .. r+CARD_IMG_ROWS-1
        img_top = r
        img_bottom = r + CARD_IMG_ROWS - 1
        cap_row = img_bottom + 1
        cap_bottom = cap_row + CAPTION_ROWS - 1

        # 이미지 캔버스 병합
        ws.merge_cells(start_row=img_top, start_column=c1, end_row=img_bottom, end_column=c2)
        set_region_border(img_top, c1, img_bottom, c2)

        # 캡션 병합
        ws.merge_cells(start_row=cap_row, start_column=c1, end_row=cap_bottom, end_column=c2)
        ws.cell(cap_row, c1, label)
        # 캡션 정렬/테두리
        for rr in range(cap_row, cap_bottom+1):
            for cc in range(c1, c2+1):
                ws.cell(rr, cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(rr, cc).border = border

        # 이미지 삽입 (4:3 패딩 + 리사이즈)
        img4x3 = enforce_aspect_pad(pil_img, 4/3).copy()
        # 엑셀 안에서 보기 좋게 축소 (대략)
        # openpyxl은 픽셀 기준 → 적절히 450x337 정도로 축소
        target_w, target_h = 450, int(450 * 3 / 4)
        img4x3.thumbnail((target_w, target_h))
        xbio = io.BytesIO()
        img4x3.save(xbio, format="PNG")
        xbio.seek(0)
        ximg = XLImage(xbio)
        # 카드 좌상단 셀에 삽입
        anchor = f"{get_column_letter(c1)}{img_top}"
        ws.add_image(ximg, anchor)

        # 다음 카드 위치 계산 (3열)
        cgi += 1
        if cgi >= 3:
            cgi = 0
            # 다음 행 블록으로 이동
            r = cap_bottom + 1 + 1  # 캡션 아래 한 줄 띄우기

    # 바이트로 저장
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# ───────────────────────────────
# 세션 상태
# ───────────────────────────────
if "photos" not in st.session_state:
    st.session_state.photos = [{"id": str(uuid.uuid4()), "choice": "장비납품", "custom": "", "checked": False, "img": None}]
if "out_ready" not in st.session_state:
    st.session_state.out_ready = False

# 옵션
CASCADE_OPTIONS = ["장비납품", "급탕모듈러설치", "난방모듈러설치", "하부배관", "LLH시공",
                   "연도시공", "외부연도마감", "드레인호스", "NCC판넬", "완료사진", "직접입력"]
VENT_OPTIONS = ["직접입력"]

# ───────────────────────────────
# 상단 UI
# ───────────────────────────────
mode = st.radio("양식 선택", ["캐스케이드", "환기"], horizontal=True)
options = CASCADE_OPTIONS if mode == "캐스케이드" else VENT_OPTIONS
site_addr = st.text_input("현장 주소", "")

st.caption("행 구성: **[체크박스]  [항목] [드롭다운]**  → (아래) 사진 등록/미리보기")

# ───────────────────────────────
# 항목 UI (한 줄: 체크박스 · '항목' 라벨 · 드롭다운  /  아래 줄: 사진 업로드 + 미리보기)
# ───────────────────────────────
for p in st.session_state.photos:
    # 한 줄
    row = st.columns([0.5, 0.7, 3.5])
    with row[0]:
        p["checked"] = st.checkbox("", key=f"chk_{p['id']}", value=p.get("checked", False))
    with row[1]:
        st.markdown("**항목**")
    with row[2]:
        p["choice"] = st.selectbox(label="", options=options, key=f"sel_{p['id']}")
        if p["choice"] == "직접입력":
            p["custom"] = st.text_input("직접입력", value=p.get("custom",""), key=f"custom_{p['id']}")
    # 아래 줄: 사진 업로드
    up = st.file_uploader("사진 등록", type=["jpg", "jpeg", "png"], key=f"up_{p['id']}")
    if up:
        p["img"] = Image.open(up).convert("RGB")
    if p.get("img"):
        st.image(p["img"], use_container_width=True)

# ───────────────────────────────
# 버튼: 추가 / 선택 삭제 / (PDF·Excel) 생성
# ───────────────────────────────
c1, c2, c3 = st.columns([1,1,2])
with c1:
    if st.button("➕ 항목 추가"):
        st.session_state.photos.append({"id": str(uuid.uuid4()), "choice": options[0], "custom": "", "checked": False, "img": None})
with c2:
    if st.button("🗑 선택 삭제"):
        st.session_state.photos = [p for p in st.session_state.photos if not p["checked"]]
        for p in st.session_state.photos:
            p["checked"] = False
with c3:
    if st.button("📄 PDF·Excel 생성", type="primary"):
        valid = [( (p["custom"].strip() if (p["choice"]=="직접입력" and p.get("custom","").strip()) else p["choice"]), p["img"])
                 for p in st.session_state.photos if p.get("img") is not None]
        if not valid:
            st.warning("📸 사진이 등록된 항목이 없습니다.")
        else:
            doc_title = f"{mode} 기성 청구 양식"
            pdf_bytes = build_pdf(doc_title, site_addr, valid)
            xlsx_bytes = build_excel_like_pdf(doc_title, site_addr, valid)
            st.session_state.out_ready = True
            st.session_state.pdf_bytes = pdf_bytes
            st.session_state.xlsx_bytes = xlsx_bytes

# ───────────────────────────────
# 다운로드 버튼 (나란히)
# ───────────────────────────────
if st.session_state.get("out_ready"):
    fname_base = f"{sanitize_filename(site_addr)}_{mode}_기성청구".strip("_")
    st.success("✅ PDF·Excel 파일이 생성되었습니다. 아래에서 내려받으세요.")
    d1, d2 = st.columns(2)
    with d1:
        st.download_button("⬇️ PDF 다운로드", st.session_state.pdf_bytes,
                           file_name=f"{fname_base}.pdf", mime="application/pdf")
    with d2:
        st.download_button("⬇️ Excel 다운로드", st.session_state.xlsx_bytes,
                           file_name=f"{fname_base}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
